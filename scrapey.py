from selenium import webdriver
from selenium.webdriver.chrome.service import Service 
from webdriver_manager.chrome import ChromeDriverManager
import time
from openpyxl import Workbook, load_workbook

# Create a new instance of the Chrome driver
driver_path = ChromeDriverManager().install()
extension_path1 ='ublock Origin.crx'
extension_path2 ='Privacy Badger.crx'
options = webdriver.ChromeOptions()
options.add_extension(extension_path1)
options.add_extension(extension_path2)
s = Service(driver_path)
driver =webdriver.Chrome(service=s,options=options)

# Navigate to the website
visit ='https://www.smartprix.com/mobiles/apple-brand/smartphone-type/exclude_out_of_stock-exclude_upcoming-stock?sort=price&asc=0'
driver.get(visit)

# Get total number of phones of this brand
var = driver.find_element("xpath",'//*[@id="app"]/main/div[1]/div[2]/div[1]/div/div[1]').text
var = int(''.join(filter(str.isdigit, var)))

# Iterate over a range of values from 1 to var
for i in range(1, var+1):
    print(i)
    # Calculate the value of var based on the value of i
    tip = (i - 1 ) // 20  
    # Navigate to the website
    driver.get(visit)
    # Find the "Load More" button and click it var times to load more phones
    if tip>0:
        for j in range(tip):
            load_button = driver.find_element("xpath", '//*[@id="app"]/main/div[1]/div[2]/div[3]')
            load_button.click()
            time.sleep(2) # Wait for the new phones to load

    # Find the phone link on the page
    try:
        phone = driver.find_element("xpath",f'//*[@id="app"]/main/div[1]/div[2]/div[2]/div[{i}]/a')
        url = phone.get_attribute("href")
        print("Visiting phone:", url)
        driver.get(url)
    except:
        print("Link not found")

    # Add any additional code to scrape data from each phone's page here
    try:
        phone_name = driver.find_element("xpath",'//*[@id="app"]/main/div[1]/div[2]/div[1]/div[1]/h1').text
    except:
        print("Element phone name not found")
        phone_name = None
    try:
        phone_price = driver.find_element("xpath",'//*[@id="app"]/main/div[1]/div[2]/div[3]/div[1]/div[1]').text
    except:
        print("Element phone price not found")
        phone_price = None
    try:
        image= driver.find_element("xpath",'//*[@id="app"]/main/div[1]/div[1]/div[1]/div[1]/div[1]/img')
        image_link = image.get_attribute("src")
    except:
        print("Element phone image not found")
        image_link = None

    #buy link 1
    try:
        buy_link1 = driver.find_element("xpath",'//*[@id="app"]/main/div[1]/div[2]/div[3]/ul/li[1]/a')
        link1_href = buy_link1.get_attribute("href")
    except:
         link1_href = None   
    #buy link 2
    try:
        buy_link2 = driver.find_element("xpath",'//*[@id="app"]/main/div[1]/div[2]/div[3]/ul/li[2]/a')
        link2_href = buy_link2.get_attribute("href")
    except:
        link2_href = None

    # Click on the "View More Specs" button and extract the data from the popup
    try:
        specs_button = driver.find_element("xpath",'//*[@id="app"]/main/div[3]/div/div[1]/div[5]/a')
        specs_button.click()
        time.sleep(2) # Wait for the popup to load
    except:
        print("Element specs button not found")

    # Add any additional code to scrape data from "View More Specs" here
    specs_data = {}

    # Loop through each popup content div and extract its data
    for i in range(1, 11):
        try:
            # Create the XPath for the current div
            popup_xpath = f'//*[@id="app"]/div[2]/div/div[2]/div/div[{i}]'
            popup_content = driver.find_element("xpath", popup_xpath)
            tables = popup_content.find_elements("xpath",'.//table')
            # Loop through each table in the div and extract its data
            for table in tables:
                rows = table.find_elements("xpath",'.//tr')
                for row in rows:
                    key = row.find_element("xpath",'.//td[1]').text
                    value = row.find_element("xpath",'.//td[2]').text
                    original_key = key
                    count = 1
                    while key in specs_data:
                        key = f"{original_key}_{count}"
                        count += 1
                    specs_data[key] = value
        except:
            print("specs xpath not found")
            specs_data = None


    # Create a new Excel workbook and sheet, or load existing file and sheet
    try:
        wb = load_workbook('AhShit.xlsx')
        ws = wb.active
        row_count = ws.max_row
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "price", "buy1", "buy2", "image"] + list(specs_data.keys()))
        row_count = 1

    # Get the column index for each key in the first row
    header_row = ws[1]
    header_dict = {cell.value: cell.column for cell in header_row}

    # Append the phone name, price, and specs to the worksheet
    row_data = [''] * len(header_dict) # Initialize row data with empty strings for all columns
    row_data[0] = phone_name
    row_data[1] = phone_price
    row_data[2] = link1_href
    row_data[3] = link2_href
    row_data[4] = image_link

    if specs_data is not None:
        for key, value in specs_data.items():
            col_index = header_dict.get(key)
            if col_index is not None:
                row_data[col_index-1] = value

    ws.append(row_data)
    row_count += 1

    # Save the Excel file
    wb.save('AhShit.xlsx')

# Close the browser window
driver.quit()



